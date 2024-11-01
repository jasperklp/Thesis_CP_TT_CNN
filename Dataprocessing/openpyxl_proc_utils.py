import openpyxl
import openpyxl.worksheet
import openpyxl.worksheet.worksheet


def print_in_first_two_cells_from_dict(sheet : openpyxl.worksheet, offset: int, measurementnr : int, row : int, key, value):
    sheet.cell(row = row, column = offset*measurementnr+1, value=f"{key}")
    sheet.cell(row = row, column = offset*measurementnr+2, value=f"{value}")

def print_line_in_minor_row(sheet : openpyxl.worksheet.worksheet.Worksheet, offset : int, measurementnr:int, row:int, args:list):
    for i in range(len(args)):
        sheet.cell(row = row, column = offset*measurementnr+i+1, value=f"{args[i]}")
